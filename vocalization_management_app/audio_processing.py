import os
import re
import uuid
import json
import logging
import tempfile
import shutil
import librosa
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from pathlib import Path
from datetime import datetime
from django.conf import settings
from django.utils.timezone import now
from scipy.io import wavfile as wav
from .models import DetectedNoiseAudioFile, Database, ProcessingLog, OriginalAudioFile, Zoo, Spectrogram
from datetime import datetime
from django.core.files.base import ContentFile
from django.core.exceptions import ObjectDoesNotExist

def ensure_directory_exists(directory):
    """Ensure that a directory exists, create if it doesn't"""
    if not os.path.exists(directory):
        os.makedirs(directory)

def organize_audio_file_by_date(audio_file_obj, recording_date):
    """
    Move an audio file to a date-based directory structure based on its recording date.
    
    """
    if not recording_date:
        return False
        
    try:
        # Create a directory structure based on the recording date
        year = recording_date.strftime('%Y')
        month = recording_date.strftime('%m')
        day = recording_date.strftime('%d')
        
        # Construct the relative path for the new location
        animal_type = original_audio.animal_type
        new_relative_dir = os.path.join('audio_files', animal_type, year, month, day)
        
        # Ensure the directory exists
        full_dir_path = os.path.join(settings.MEDIA_ROOT, new_relative_dir)
        os.makedirs(full_dir_path, exist_ok=True)
        
        # Get the filename
        filename = os.path.basename(original_audio.audio_file.name)
        
        # Construct the new path
        new_relative_path = os.path.join(new_relative_dir, filename)
        new_full_path = os.path.join(settings.MEDIA_ROOT, new_relative_path)
        
        # Get the current full path
        current_full_path = original_audio.audio_file.path
        
        # If the file is already in the correct location, no need to move it
        if os.path.normpath(current_full_path) == os.path.normpath(new_full_path):
            return True
        
        # Create a temporary file to avoid name conflicts
        temp_dir = tempfile.mkdtemp(dir=os.path.dirname(new_full_path))
        temp_path = os.path.join(temp_dir, filename)
        
        # Copy to temp location first to avoid issues with moving across filesystems
        shutil.copy2(current_full_path, temp_path)
        
        # Update the model to point to the new location
        original_audio.audio_file.name = new_relative_path
        original_audio.save()
        
        # Move from temp to final destination
        shutil.move(temp_path, new_full_path)
        
        # Clean up temp directory
        shutil.rmtree(temp_dir, ignore_errors=True)
        
        # Remove the original file if it's different from the new location
        if os.path.exists(current_full_path) and current_full_path != new_full_path:
            os.remove(current_full_path)
        
        return True
        
    except Exception as e:
        # Log the error
        ProcessingLog.objects.create(
            audio_file=original_audio,
            message=f"Error organizing file: {str(e)}",
            level="ERROR"
        )
        return False


def generate_spectrogram(audio_file_path, original_audio, spectrogram_type='mel'):
    """
    Generate and save a spectrogram from an audio file.
    
    Parameters:
    - audio_file_path: Path to the audio file
    - original_audio: OriginalAudioFile model instance
    - spectrogram_type: Type of spectrogram to generate ('mel', 'linear', or 'chroma')
    
    Returns:
    - Spectrogram model instance or None if failed
    """
    try:
        # Create a log entry for the start of spectrogram generation
        ProcessingLog.objects.create(
            audio_file=original_audio,
            message=f"Starting {spectrogram_type} spectrogram generation",
            level="INFO"
        )
        
        # Load the audio file
        try:
            # Use librosa for better handling of various audio formats
            y, sr = librosa.load(audio_file_path, sr=None)
        except Exception as e:
            # If librosa fails, try with scipy as a fallback
            ProcessingLog.objects.create(
                audio_file=original_audio,
                message=f"Librosa failed to load audio, trying scipy: {str(e)}",
                level="WARNING"
            )
            sr, y = wav.read(audio_file_path)
            # Convert to float32 for librosa compatibility if needed
            if y.dtype != np.float32:
                y = y.astype(np.float32) / np.iinfo(y.dtype).max
        
        # Generate the spectrogram based on the type
        if spectrogram_type == 'mel':
            # Mel spectrogram (good for general audio analysis)
            S = librosa.feature.melspectrogram(y=y, sr=sr, n_mels=128, 
                                            fmax=sr/2, hop_length=512)
            # Convert to dB scale
            S_dB = librosa.power_to_db(S, ref=np.max)
            title = "Mel Spectrogram"
        elif spectrogram_type == 'linear':
            # Linear spectrogram (standard STFT)
            S = np.abs(librosa.stft(y, hop_length=512))
            # Convert to dB scale
            S_dB = librosa.amplitude_to_db(S, ref=np.max)
            title = "Linear Spectrogram"
        elif spectrogram_type == 'chroma':
            # Chromagram (good for tonal content)
            S_dB = librosa.feature.chroma_stft(y=y, sr=sr, hop_length=512)
            title = "Chromagram"
        else:
            # Default to mel spectrogram
            S = librosa.feature.melspectrogram(y=y, sr=sr)
            S_dB = librosa.power_to_db(S, ref=np.max)
            title = "Mel Spectrogram"
        
        # Create a figure for the spectrogram
        plt.figure(figsize=(10, 4))
        
        # Plot the spectrogram
        librosa.display.specshow(S_dB, sr=sr, x_axis='time', y_axis='mel' if spectrogram_type != 'chroma' else 'chroma')
        plt.colorbar(format='%+2.0f dB')
        plt.title(title)
        
        # Save the figure to a temporary file
        with tempfile.NamedTemporaryFile(suffix='.png', delete=False) as temp_file:
            temp_path = temp_file.name
            plt.savefig(temp_path, bbox_inches='tight', dpi=300)
            plt.close()
        
        # Create a relative path for the spectrogram
        filename = os.path.basename(original_audio.audio_file.name)
        base_filename = os.path.splitext(filename)[0]
        spectrogram_filename = f"{base_filename}_{spectrogram_type}_spectrogram.png"
        
        # Create a directory structure based on the recording date
        if original_audio.recording_date:
            year = original_audio.recording_date.strftime('%Y')
            month = original_audio.recording_date.strftime('%m')
            day = original_audio.recording_date.strftime('%d')
        else:
            # Use upload date as fallback
            year = original_audio.upload_date.strftime('%Y')
            month = original_audio.upload_date.strftime('%m')
            day = original_audio.upload_date.strftime('%d')
        
        # Construct the relative path for the spectrogram
        spectrogram_relative_dir = os.path.join('spectrograms', original_audio.animal_type, year, month, day)
        
        # Ensure the directory exists
        spectrogram_full_dir = os.path.join(settings.MEDIA_ROOT, spectrogram_relative_dir)
        os.makedirs(spectrogram_full_dir, exist_ok=True)
        
        # Construct the full path for the spectrogram
        spectrogram_relative_path = os.path.join(spectrogram_relative_dir, spectrogram_filename)
        spectrogram_full_path = os.path.join(settings.MEDIA_ROOT, spectrogram_relative_path)
        
        # Move the temporary file to the final location
        shutil.move(temp_path, spectrogram_full_path)
        
        # Create or update the Spectrogram model
        spectrogram, created = Spectrogram.objects.update_or_create(
            audio_file=original_audio,
            spectrogram_type=spectrogram_type,
            defaults={
                'spectrogram_file': spectrogram_relative_path,
                'generated_date': now()
            }
        )
        
        # Log successful spectrogram generation
        ProcessingLog.objects.create(
            audio_file=original_audio,
            message=f"{spectrogram_type.capitalize()} spectrogram generated successfully",
            level="SUCCESS"
        )
        
        return spectrogram
        
    except Exception as e:
        # Log error in spectrogram generation
        ProcessingLog.objects.create(
            audio_file=original_audio,
            message=f"Error generating {spectrogram_type} spectrogram: {str(e)}",
            level="ERROR"
        )
        return None

def parse_audio_filename(filename):
    """
    Parse audio filenames in various formats, with primary support for:
    - Standard format: SMM07257_20230201_171502.wav
    - With animal prefix: amur_leopard_SMM07257_20230201_171502.wav
    
    But also handles many other formats including:
    - Date-first: 20230201_171502_SMM07257.wav
    - Combined date-time: SMM07257_20230201171502.wav
    - Dash separated: SMM07257-20230201-171502.wav
    - With additional segments: SMM07257_20230201_171502_additional_info.wav
    - ISO format dates: SMM07257_2023-02-01_17-15-02.wav
    - Natural language dates: SMM07257_Feb-01-2023_5-15-02-PM.wav
    
    Returns:
        tuple: (device_info, recording_datetime)
            device_info: dict containing:
                - device_type: str (e.g., 'SMM' for Song Meter Micro)
                - device_id: str (e.g., '07257')
                - full_device_id: str (e.g., 'SMM07257')
            recording_datetime: datetime object or None if date/time couldn't be parsed
    """
    import re
    from datetime import datetime
    from django.utils import timezone
    import logging
    
    logger = logging.getLogger(__name__)
    
    # Remove file extension if present
    base_name = os.path.splitext(filename)[0]
    
    # Initialize default return values
    default_device_info = {
        'device_type': 'UNK',  # Unknown device type
        'device_id': '00000',  # Default device ID
        'full_device_id': 'UNK00000'  # Default full device ID
    }
    
    # Get current time as fallback
    current_time = timezone.now()
    
    try:
        # Step 1: Remove animal type prefixes if present
        animal_types = ['amur_leopard', 'amur_tiger']
        for animal_type in animal_types:
            prefix = f"{animal_type}_"
            if base_name.startswith(prefix):
                base_name = base_name[len(prefix):]
                break
        
        # Step 2: Split the filename into parts using common separators
        # First try underscore, then dash, then space
        if '_' in base_name:
            parts = base_name.split('_')
        elif '-' in base_name:
            parts = base_name.split('-')
        else:
            # Try to find patterns in the filename
            parts = re.findall(r'[A-Za-z]+|\d+', base_name)
        
        # Step 3: Extract device ID, date, and time components
        device_id = None
        date_str = None
        time_str = None
        
        # Pattern matching for device ID (typically starts with letters followed by numbers)
        device_pattern = re.compile(r'^([A-Za-z]+)(\d+)$')
        
        # Date patterns (YYYYMMDD, YYYY-MM-DD, etc.)
        date_patterns = [
            # YYYYMMDD
            re.compile(r'^(\d{4})(\d{2})(\d{2})$'),
            # YYYY-MM-DD or YYYY/MM/DD
            re.compile(r'^(\d{4})[-/](\d{1,2})[-/](\d{1,2})$'),
            # MM-DD-YYYY or MM/DD/YYYY
            re.compile(r'^(\d{1,2})[-/](\d{1,2})[-/](\d{4})$'),
            # Month name formats: Jan-01-2023, January-01-2023, etc.
            re.compile(r'^([A-Za-z]{3,9})[-\s](\d{1,2})[-\s](\d{4})$'),
            re.compile(r'^(\d{1,2})[-\s]([A-Za-z]{3,9})[-\s](\d{4})$'),
            re.compile(r'^(\d{4})[-\s]([A-Za-z]{3,9})[-\s](\d{1,2})$')
        ]
        
        # Time patterns (HHMMSS, HH:MM:SS, etc.)
        time_patterns = [
            # HHMMSS
            re.compile(r'^(\d{2})(\d{2})(\d{2})$'),
            # HH-MM-SS or HH:MM:SS
            re.compile(r'^(\d{1,2})[-:](\d{1,2})[-:](\d{1,2})$'),
            # With AM/PM: HH-MM-SS-AM, HH:MM:SS AM, etc.
            re.compile(r'^(\d{1,2})[-:](\d{1,2})[-:](\d{1,2})[-\s]?([AaPp][Mm])$')
        ]
        
        # Month name mapping
        month_names = {
            'jan': 1, 'january': 1,
            'feb': 2, 'february': 2,
            'mar': 3, 'march': 3,
            'apr': 4, 'april': 4,
            'may': 5, 'may': 5,
            'jun': 6, 'june': 6,
            'jul': 7, 'july': 7,
            'aug': 8, 'august': 8,
            'sep': 9, 'september': 9,
            'oct': 10, 'october': 10,
            'nov': 11, 'november': 11,
            'dec': 12, 'december': 12
        }
        
        # Step 4: Analyze each part to identify device ID, date, and time
        for part in parts:
            # Skip empty parts
            if not part:
                continue
                
            # Try to match device ID pattern
            if device_id is None:
                device_match = device_pattern.match(part)
                if device_match:
                    device_type = device_match.group(1)
                    unit_number = device_match.group(2)
                    device_id = part
                    continue
            
            # Try to match date patterns
            if date_str is None:
                for pattern in date_patterns:
                    date_match = pattern.match(part)
                    if date_match:
                        date_str = part
                        break
                        
                # Check for combined date-time format (YYYYMMDDHHMMSS)
                if date_str is None and len(part) >= 14 and part.isdigit():
                    date_str = part[:8]
                    time_str = part[8:14]
                    break
                    
                if date_str is not None:
                    continue
            
            # Try to match time patterns
            if time_str is None:
                for pattern in time_patterns:
                    time_match = pattern.match(part)
                    if time_match:
                        time_str = part
                        break
                        
                if time_str is not None:
                    continue
            
            # If we've already found all components, we can stop
            if device_id is not None and date_str is not None and time_str is not None:
                break
        
        # Step 5: If we still don't have all components, try alternative approaches
        
        # If no device ID found, use the first part or a default
        if device_id is None and len(parts) > 0:
            device_id = parts[0]
            
        # If no date found, check for date-like patterns in the filename
        if date_str is None:
            # Look for 8-digit sequences that could be dates
            date_matches = re.findall(r'\d{8}', base_name)
            if date_matches:
                date_str = date_matches[0]
                
        # If no time found, check for 6-digit sequences that could be times
        if time_str is None:
            # Look for 6-digit sequences that could be times
            time_matches = re.findall(r'\d{6}', base_name)
            if time_matches and (not date_str or time_matches[0] != date_str[:6]):
                time_str = time_matches[0]
        
        # Step 6: Parse the device ID
        device_info = default_device_info.copy()
        if device_id:
            # Extract device type (letters) and unit number (digits)
            device_type_match = re.match(r'^([A-Za-z]+)', device_id)
            device_id_match = re.match(r'^[A-Za-z]*(\d+)', device_id)
            
            if device_type_match:
                device_info['device_type'] = device_type_match.group(1)
            
            if device_id_match:
                device_info['device_id'] = device_id_match.group(1)
            
            device_info['full_device_id'] = device_id
        
        # Step 7: Parse the date and time
        recording_datetime = None
        
        if date_str:
            year, month, day = None, None, None
            
            # Try different date formats
            for pattern in date_patterns:
                match = pattern.match(date_str)
                if match:
                    groups = match.groups()
                    
                    # YYYYMMDD format
                    if len(groups) == 3 and len(groups[0]) == 4 and groups[0].isdigit():
                        year = int(groups[0])
                        if groups[1].isdigit():
                            month = int(groups[1])
                        else:
                            # Handle month names
                            month_name = groups[1].lower()
                            month = month_names.get(month_name, 1)
                        day = int(groups[2])
                    
                    # MM-DD-YYYY format
                    elif len(groups) == 3 and len(groups[2]) == 4 and groups[2].isdigit():
                        year = int(groups[2])
                        if groups[0].isdigit():
                            month = int(groups[0])
                        else:
                            # Handle month names
                            month_name = groups[0].lower()
                            month = month_names.get(month_name, 1)
                        day = int(groups[1])
                    
                    break
            
            # If no pattern matched, try direct parsing for YYYYMMDD
            if year is None and len(date_str) >= 8 and date_str[:8].isdigit():
                year = int(date_str[:4])
                month = int(date_str[4:6])
                day = int(date_str[6:8])
            
            # Parse time if available
            hour, minute, second = 0, 0, 0
            if time_str:
                # Try different time formats
                for pattern in time_patterns:
                    match = pattern.match(time_str)
                    if match:
                        groups = match.groups()
                        hour = int(groups[0])
                        minute = int(groups[1])
                        second = int(groups[2])
                        
                        # Handle AM/PM if present
                        if len(groups) > 3 and groups[3] and groups[3].lower() == 'pm' and hour < 12:
                            hour += 12
                        elif len(groups) > 3 and groups[3] and groups[3].lower() == 'am' and hour == 12:
                            hour = 0
                            
                        break
                
                # If no pattern matched, try direct parsing for HHMMSS
                if hour == 0 and len(time_str) >= 6 and time_str[:6].isdigit():
                    hour = int(time_str[:2])
                    minute = int(time_str[2:4])
                    second = int(time_str[4:6])
            
            # Create datetime object if we have valid date components
            if year and month and day:
                try:
                    naive_datetime = datetime(year, month, day, hour, minute, second)
                    recording_datetime = timezone.make_aware(naive_datetime, timezone.get_current_timezone())
                except (ValueError, OverflowError) as e:
                    logger.warning(f"Invalid date/time values in filename {filename}: {e}")
                    recording_datetime = current_time
        
        # If we couldn't parse a date, use current time as fallback
        if recording_datetime is None:
            recording_datetime = current_time
            logger.warning(f"Could not parse date/time from filename {filename}, using current time as fallback")
        
        return device_info, recording_datetime
        
    except Exception as e:
        logger.warning(f"Error parsing filename {filename}: {e}")
        # Instead of raising an exception, return default values
        # This allows processing to continue even with unexpected filename formats
        return default_device_info, current_time

def update_audio_metadata(file_path, original_audio):
    """
    Update the metadata of the audio file based on its filename and content
    This function extracts metadata from the filename and audio file itself
    """
    try:
        # Create a log entry for the start of metadata update
        ProcessingLog.objects.create(
            audio_file=original_audio,
            message="Starting metadata extraction",
            level="INFO"
        )
        
        # Get the filename from the path
        filename = os.path.basename(file_path)
        
        # Get recording date from filename (will use fallback if parsing fails)
        recording_datetime = None
        try:
            device_info, recording_datetime = parse_audio_filename(filename)
            
            # Update the recording date in the database
            if recording_datetime:
                original_audio.recording_date = recording_datetime
                original_audio.save()
                
                # Log successful metadata extraction
                ProcessingLog.objects.create(
                    audio_file=original_audio,
                    message=f"Successfully extracted metadata. Recording date: {recording_datetime}",
                    level="SUCCESS"
                )
                
                # Move the file to its date-based directory structure
                if organize_audio_file_by_date(original_audio, recording_datetime):
                    # File was moved successfully, update the path for subsequent operations
                    file_path = original_audio.audio_file.path
                    
                    # Log the new path
                    ProcessingLog.objects.create(
                        audio_file=original_audio,
                        message=f"File moved to: {original_audio.audio_file.name}",
                        level="INFO"
                    )
        except Exception as e:
            # Log error in metadata extraction but continue processing
            ProcessingLog.objects.create(
                audio_file=original_audio,
                message=f"Error parsing filename: {str(e)}. Using upload date as fallback.",
                level="WARNING"
            )
            
            # Use upload date as fallback if we couldn't parse the filename
            if not recording_datetime or not original_audio.recording_date:
                original_audio.recording_date = original_audio.upload_date
                original_audio.save()
                
                # Log fallback to upload date
                ProcessingLog.objects.create(
                    audio_file=original_audio,
                    message=f"Using upload date as recording date: {original_audio.upload_date}",
                    level="INFO"
                )
        
        # Extract audio properties from the file itself
        try:
            # Get file size in MB
            file_size_bytes = os.path.getsize(file_path)
            file_size_mb = file_size_bytes / (1024 * 1024)  # Convert to MB
            
            # Update file size in the database
            original_audio.file_size_mb = file_size_mb
            
            # Extract audio duration and sample rate using scipy.io.wavfile
            sample_rate, audio_data = wav.read(file_path)
            
            # Calculate duration in seconds
            duration_seconds = len(audio_data) / sample_rate
            
            # Format duration as HH:MM:SS
            hours, remainder = divmod(int(duration_seconds), 3600)
            minutes, seconds = divmod(remainder, 60)
            duration_str = f"{hours:02d}:{minutes:02d}:{seconds:02d}"
            
            # Update audio properties in the database
            original_audio.sample_rate = sample_rate
            original_audio.duration_seconds = duration_seconds
            original_audio.duration = duration_str
            
            # Log audio properties
            ProcessingLog.objects.create(
                audio_file=original_audio,
                message=f"Audio properties: {sample_rate}Hz, {duration_str} ({duration_seconds:.2f}s), {file_size_mb:.2f}MB",
                level="INFO"
            )
        except Exception as e:
            # Log error in audio property extraction but continue processing
            ProcessingLog.objects.create(
                audio_file=original_audio,
                message=f"Error extracting audio properties: {str(e)}",
                level="WARNING"
            )
        
        # Save all changes to the database
        original_audio.save()
        
        # Create database entry with Pending status if it doesn't exist
        db_entry, created = Database.objects.get_or_create(
            audio_file=original_audio,
            defaults={'status': 'Pending'}
        )
        
        # Log successful metadata update
        ProcessingLog.objects.create(
            audio_file=original_audio,
            message="Metadata update completed successfully",
            level="SUCCESS"
        )
        
        return True
        
    except Exception as e:
        # Log error in metadata update
        ProcessingLog.objects.create(
            audio_file=original_audio,
            message=f"Error updating metadata: {str(e)}",
            level="ERROR"
        )
        return False

def seconds_to_timestamp(seconds):
    """
    Converts seconds to a timestamp in HH:MM:SS.SS format, with seconds rounded to two decimal places.
    
    Parameters:
    - seconds (float): The number of seconds to convert.
    
    Returns:
    - str: The timestamp in HH:MM:SS.SS format.
    """
    hours, remainder = divmod(seconds, 3600)
    minutes, seconds = divmod(remainder, 60)
    return f"{int(hours):02d}:{int(minutes):02d}:{seconds:06.2f}"


def detect_saw_calls(audio_data, sample_rate, animal_type=None):
    """
    Detects saw calls in the audio data using STFT analysis.
    
    Parameters:
    - audio_data (numpy.ndarray): The audio data.
    - sample_rate (int): The sample rate of the audio data.
    - animal_type (str, optional): The type of animal to use parameters for. If None, uses default parameters.
    
    Returns:
    - list: A list of dictionaries containing information about detected saw calls:
            [{'start': start_time_str, 'end': end_time_str, 'magnitude': mag, 'frequency': freq, 'impulse_count': count}]
    """
    # Import here to avoid circular imports
    from .models import AnimalDetectionParameters
    
    # Get parameters from the database based on animal type or use default
    try:
        if animal_type:
            params = AnimalDetectionParameters.objects.filter(slug=animal_type).first()
        else:
            params = AnimalDetectionParameters.objects.filter(is_default=True).first()
            
        # If no parameters found, use default hardcoded values
        if params:
            min_mag = params.min_magnitude
            max_mag = params.max_magnitude
            min_freq = params.min_frequency
            max_freq = params.max_frequency
            segment_duration = params.segment_duration
            time_threshold = params.time_threshold
            min_impulse_count = params.min_impulse_count
        else:
            # Fallback to hardcoded defaults (Amur Leopard)
            min_mag = 3500
            max_mag = 10000
            min_freq = 15
            max_freq = 300
            segment_duration = 0.1
            time_threshold = 5
            min_impulse_count = 3
    except Exception as e:
        # Log the error and use defaults
        print(f"Error loading detection parameters: {e}")
        min_mag = 3500
        max_mag = 10000
        min_freq = 15
        max_freq = 300
        segment_duration = 0.1
        time_threshold = 5
        min_impulse_count = 3
    
    # Convert audio data to float32 if not already
    audio_data = audio_data.astype(np.float32)
    
    # Remove DC offset by subtracting the mean
    audio_data -= np.mean(audio_data)
    
    # Compute the STFT with a specified segment duration
    nperseg = int(segment_duration * sample_rate)
    frequencies, times, Zxx = stft(audio_data, fs=sample_rate, nperseg=nperseg)
    magnitude = np.abs(Zxx)  # Magnitude of the STFT result
    
    # Initialize variables for event detection
    last_event_time_seconds = None
    event_data = []
    
    # Iterate over each time frame
    for time_idx in range(magnitude.shape[1]):
        # Filter for magnitudes above the threshold
        magnitudes_at_time = magnitude[:, time_idx]
        valid_indices = np.where(np.logical_and(magnitudes_at_time < max_mag, magnitudes_at_time > min_mag))[0]
        
        if valid_indices.size == 0:
            continue  # Skip if no magnitudes exceed the threshold
        
        event_time_seconds = times[time_idx]
        # Format with higher precision (2 decimal places for seconds)
        event_time_str = seconds_to_timestamp(event_time_seconds)
        
        # Extract valid frequencies and magnitudes
        valid_frequencies = frequencies[valid_indices]
        valid_magnitudes = magnitudes_at_time[valid_indices]
        
        # Filter frequencies between min_freq and max_freq Hz
        freq_mask = (valid_frequencies < max_freq) & (valid_frequencies > min_freq)
        valid_frequencies = valid_frequencies[freq_mask]
        valid_magnitudes = valid_magnitudes[freq_mask]
        
        if len(valid_frequencies) == 0:
            continue  # Skip if no frequencies in the valid range
        
        # Process events
        for freq, mag in zip(valid_frequencies, valid_magnitudes):
            if last_event_time_seconds is None:
                # First event, add to dataset
                event_data.append({
                    'start': event_time_str,
                    'end': event_time_str,
                    'start_seconds': event_time_seconds,  # Store raw seconds for precise calculations
                    'end_seconds': event_time_seconds,    # Store raw seconds for precise calculations
                    'magnitude': float(mag),
                    'frequency': float(freq),
                    'impulse_count': 1
                })
                last_event_time_seconds = event_time_seconds
            else:
                # Check if the event is within the time threshold and more than 0.1 seconds later
                time_diff = event_time_seconds - last_event_time_seconds
                
                if time_diff <= time_threshold and time_diff > 0.1:  # Events within threshold are merged
                    # Extend the duration of the last event
                    event_data[-1]['end'] = event_time_str
                    event_data[-1]['end_seconds'] = event_time_seconds
                    # Increase count of impulses
                    event_data[-1]['impulse_count'] += 1
                    # Update magnitude and frequency if higher
                    if mag > event_data[-1]['magnitude']:
                        event_data[-1]['magnitude'] = float(mag)
                        event_data[-1]['frequency'] = float(freq)
                elif time_diff > time_threshold:
                    # Add new event
                    event_data.append({
                        'start': event_time_str,
                        'end': event_time_str,
                        'start_seconds': event_time_seconds,
                        'end_seconds': event_time_seconds,
                        'magnitude': float(mag),
                        'frequency': float(freq),
                        'impulse_count': 1
                    })
                
                # Update last event time
                last_event_time_seconds = event_time_seconds
    
    # Filter out events with less than 3 impulses (likely false positives)
    filtered_events = [event for event in event_data if event['impulse_count'] >= 3]
    
    return filtered_events


def generate_excel_report(original_audio, saw_calls):
    """
    Generate an Excel report for the detected saw calls.
    
    Parameters:
    - original_audio: OriginalAudioFile instance
    - saw_calls: List of dictionaries containing saw call data

    Returns:
    - Path to the saved Excel file
    """
    try:
        # Use the globally imported pandas
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter

        # Create a directory for Excel reports if it doesn't exist
        excel_dir = os.path.join(settings.MEDIA_ROOT, 'excel_reports')
        if not os.path.exists(excel_dir):
            os.makedirs(excel_dir)

        # Generate a filename based on the audio file name
        base_name = os.path.splitext(original_audio.audio_file_name)[0]
        excel_filename = f"{base_name}_report.xlsx"
        excel_path = os.path.join(excel_dir, excel_filename)

        # Create a pandas DataFrame from the saw calls data
        data = []
        for call in saw_calls:
            # Extract date from filename if available
            recording_date = original_audio.recording_date.strftime('%Y-%m-%d') if original_audio.recording_date else 'Unknown'

            # Calculate duration
            start_seconds = call.get('start_seconds', 0)
            end_seconds = call.get('end_seconds', 0)
            duration = end_seconds - start_seconds if end_seconds > start_seconds else 0
            duration_str = seconds_to_timestamp(duration)

            # Add row data
            data.append({
                'File Name': original_audio.audio_file_name,
                'Recording Date': recording_date,
                'Animal Type': original_audio.animal_type,
                'Start Time': call.get('start', '00:00:00'),
                'End Time': call.get('end', '00:00:00'),
                'Duration': duration_str,
                'Frequency (Hz)': call.get('frequency', 0),
                'Magnitude': call.get('magnitude', 0),
                'Impulse Count': call.get('impulse_count', 0)
            })

        # Create DataFrame
        df = pd.DataFrame(data)

        # Create a workbook and select the active worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Saw Calls Report"

        # Add a title
        ws['A1'] = f"Saw Calls Report for {original_audio.audio_file_name}"
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:I1')
        ws['A1'].alignment = Alignment(horizontal='center')

        # Add metadata
        recording_date_str = original_audio.recording_date.strftime('%Y-%m-%d') if original_audio.recording_date else 'Unknown'
        ws['A2'] = f"Recording Date: {recording_date_str}"
        ws['A3'] = f"Animal Type: {original_audio.animal_type}"
        ws['A4'] = f"Total Saw Calls Detected: {len(saw_calls)}"

        # Add headers at row 6
        headers = list(df.columns)
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=6, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')

        # Add data starting at row 7
        for row_num, row_data in enumerate(df.values, 7):
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.alignment = Alignment(horizontal='center')

        # Auto-adjust column widths
        for col_num, _ in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            # Set a minimum width
            ws.column_dimensions[col_letter].width = 15

        # Save the workbook
        wb.save(excel_path)

        return excel_path

        # Create a BytesIO object to save the Excel file
        excel_file = io.BytesIO()
        
        # Create ExcelWriter object with the BytesIO object
        with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
            metadata_df.to_excel(writer, sheet_name='Metadata', index=False)
            df.to_excel(writer, sheet_name='Saw Calls', index=False)
            
            # Auto-adjust column widths
            for sheet_name in writer.sheets:
                worksheet = writer.sheets[sheet_name]
                for i, col in enumerate(worksheet.columns):
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    adjusted_width = (max_length + 2)
                    worksheet.column_dimensions[column].width = adjusted_width
        
        # Reset the pointer to the beginning of the BytesIO object
        excel_file.seek(0)
        
        # Generate a filename for the Excel file
        excel_filename = f"{os.path.splitext(original_audio.audio_file_name)[0]}_analysis.xlsx"
        
        # Save the Excel file to the OriginalAudioFile model
        original_audio.analysis_excel.save(excel_filename, ContentFile(excel_file.read()), save=True)
        
        # Log successful Excel generation
        ProcessingLog.objects.create(
            audio_file=original_audio,
            message=f"Generated Excel report: {excel_filename}",
            level="SUCCESS"
        )
        
        return original_audio.analysis_excel.path
        
    except Exception as e:
        # Log error in Excel generation
        ProcessingLog.objects.create(
            audio_file=original_audio,
            message=f"Error generating Excel report: {str(e)}",
            level="ERROR"
        )
        return None

    """
    Process the uploaded audio file and store saw call timeframes.
    Uses improved STFT-based detection to accurately identify and log saw calls.
    """
    temp_files = []  # Track temporary files for cleanup
    
    try:
        # Check if this file has already been processed
        db_entry = Database.objects.filter(audio_file=original_audio).first()
        
        if db_entry and db_entry.status == 'Processed':
            # Log that file is already processed
            ProcessingLog.objects.create(
                audio_file=original_audio,
                message="File already processed. Skipping processing.",
                level="INFO"
            )
            return True
        
        # Update database entry to Processing status
        if db_entry:
            db_entry.status = 'Processing'
            db_entry.processing_start_time = now()
            db_entry.save()
        else:
            db_entry = Database.objects.create(
                audio_file=original_audio,
                status='Processing',
                processing_start_time=now()
            )
        
        # Log processing start
        ProcessingLog.objects.create(
            audio_file=original_audio,
            message="Starting audio processing",
            level="INFO"
        )
        
        # Load the audio file
        try:
            # First try using scipy.io.wavfile for better compatibility with various WAV formats
            try:
                ProcessingLog.objects.create(
                    audio_file=original_audio,
                    message="Attempting to load audio with scipy.io.wavfile",
                    level="INFO"
                )
                
                # Suppress warnings when loading WAV files
                import warnings
                with warnings.catch_warnings():
                    warnings.simplefilter("ignore")
                    sample_rate, audio_data = wav.read(file_path)
                
                # Convert to mono if stereo
                if isinstance(audio_data, np.ndarray) and len(audio_data.shape) > 1:
                    audio_data = np.mean(audio_data, axis=1)
                
                ProcessingLog.objects.create(
                    audio_file=original_audio,
                    message=f"Loaded audio file with scipy.io.wavfile: {sample_rate}Hz, {len(audio_data)/sample_rate:.2f}s",
                    level="SUCCESS"
                )
            except Exception as e:
                # If scipy fails, try librosa
                ProcessingLog.objects.create(
                    audio_file=original_audio,
                    message=f"scipy.io.wavfile failed: {str(e)}. Trying librosa...",
                    level="WARNING"
                )
                
                audio_data, sample_rate = librosa.load(file_path, sr=None, mono=True)
                
                ProcessingLog.objects.create(
                    audio_file=original_audio,
                    message=f"Loaded audio file with librosa: {sample_rate}Hz, {len(audio_data)/sample_rate:.2f}s",
                    level="SUCCESS"
                )
            
            # Update file metadata if not already set
            if not original_audio.duration_seconds:
                duration_seconds = len(audio_data) / sample_rate
                original_audio.duration_seconds = duration_seconds
                original_audio.duration = seconds_to_timestamp(duration_seconds)
                original_audio.sample_rate = sample_rate
                original_audio.save()
                
                ProcessingLog.objects.create(
                    audio_file=original_audio,
                    message=f"Updated file metadata: Duration={original_audio.duration}, Sample Rate={sample_rate}Hz",
                    level="INFO"
                )
            
            # Detect saw calls using STFT analysis
            ProcessingLog.objects.create(
                audio_file=original_audio,
                message="Starting saw call detection with STFT analysis",
                level="INFO"
            )
            
            saw_calls = detect_saw_calls(audio_data, sample_rate)
            
            # Filter out calls with less than 3 impulses (likely false positives)
            filtered_saw_calls = [call for call in saw_calls if call['impulse_count'] >= 3]
            
            ProcessingLog.objects.create(
                audio_file=original_audio,
                message=f"Detected {len(filtered_saw_calls)} saw calls after filtering (minimum 3 impulses required)",
                level="SUCCESS"
            )
            
        except Exception as e:
            # Log error in saw call detection
            ProcessingLog.objects.create(
                audio_file=original_audio,
                message=f"Error detecting saw calls: {str(e)}",
                level="ERROR"
            )
            
            # Update database entry to Failed status
            db_entry.status = 'Failed'
            db_entry.processing_end_time = now()
            db_entry.save()
            return False
        
        # Extract date from filename if available
        file_date = None
        try:
            device_info, recording_datetime = parse_audio_filename(original_audio.audio_file_name)
            # Update the recording date in the database if not already set
            if not original_audio.recording_date:
                original_audio.recording_date = recording_datetime
                original_audio.save()
            file_date = recording_datetime.date()
        except Exception as e:
            # Log warning but continue processing
            ProcessingLog.objects.create(
                audio_file=original_audio,
                message=f"Warning: Could not extract date from filename: {str(e)}",
                level="WARNING"
            )
        
        # Store saw call timeframes
        saw_count = 0
        for saw_call in filtered_saw_calls:
            try:
                # Convert timestamp strings to time objects for database storage
                start_seconds = saw_call['start_seconds']
                end_seconds = saw_call['end_seconds']
                
                # Calculate hours, minutes, seconds for start time
                start_hours, start_remainder = divmod(int(start_seconds), 3600)
                start_minutes, start_seconds_int = divmod(start_remainder, 60)
                start_microseconds = int((start_seconds - int(start_seconds)) * 1000000)
                
                # Calculate hours, minutes, seconds for end time
                end_hours, end_remainder = divmod(int(end_seconds), 3600)
                end_minutes, end_seconds_int = divmod(end_remainder, 60)
                end_microseconds = int((end_seconds - int(end_seconds)) * 1000000)
                
                # Create time objects
                from datetime import time
                start_time_obj = time(start_hours, start_minutes, start_seconds_int, start_microseconds)
                end_time_obj = time(end_hours, end_minutes, end_seconds_int, end_microseconds)
                
                # Create detected noise entry
                DetectedNoiseAudioFile.objects.create(
                    original_file=original_audio,
                    detected_noise_file_path="",  # We're not creating actual files
                    start_time=start_time_obj,
                    end_time=end_time_obj,
                    saw_count=saw_call['impulse_count'],  # Number of impulses in this call
                    saw_call_count=1,  # Each entry represents one saw call
                    file_size_mb=0.0,  # No actual file
                    frequency=saw_call['frequency'],
                    magnitude=saw_call['magnitude']
                )
                saw_count += 1
                
                # Log individual saw call detection with precise timestamps
                ProcessingLog.objects.create(
                    audio_file=original_audio,
                    message=f"Detected saw call: Start={saw_call['start']}, End={saw_call['end']}, Duration={(end_seconds-start_seconds):.2f}s, Impulses={saw_call['impulse_count']}, Freq={saw_call['frequency']:.2f}Hz, Mag={saw_call['magnitude']:.2f}",
                    level="INFO"
                )
            except Exception as e:
                ProcessingLog.objects.create(
                    audio_file=original_audio,
                    message=f"Error storing saw call data: {str(e)}",
                    level="WARNING"
                )
        
        # If no saw calls were detected, log this explicitly
        if saw_count == 0:
            ProcessingLog.objects.create(
                audio_file=original_audio,
                message="No saw calls detected in this audio file",
                level="INFO"
            )
        else:
            # Log successful storage of saw calls
            ProcessingLog.objects.create(
                audio_file=original_audio,
                message=f"Successfully stored {saw_count} saw call timeframes",
                level="SUCCESS"
            )
        
        # Generate Excel report
        try:
            excel_path = generate_excel_report(original_audio, saw_calls)
            temp_files.append(excel_path)  # Track the Excel file for potential cleanup
            
            # Log Excel report generation
            ProcessingLog.objects.create(
                audio_file=original_audio,
                message=f"Excel report generated successfully: {excel_path}",
                level="INFO"
            )
        except Exception as e:
            # Log error but continue processing
            ProcessingLog.objects.create(
                audio_file=original_audio,
                message=f"Error generating Excel report: {str(e)}",
                level="ERROR"
            )
        
        # Update database entry to Processed status
        db_entry.status = 'Processed'
        db_entry.processing_end_time = now()
        db_entry.save()
        
        # Log successful processing completion
        ProcessingLog.objects.create(
            audio_file=original_audio,
            message="Audio processing completed successfully",
            level="SUCCESS"
        )
        
        return True
        
    except Exception as e:
        # Log the error
        ProcessingLog.objects.create(
            audio_file=original_audio,
            message=f"Error processing audio file: {str(e)}",
            level="ERROR"
        )
        
        # Update database entry to Failed status
        db_entry = Database.objects.filter(audio_file=original_audio).first()
        if db_entry:
            db_entry.status = 'Failed'
            db_entry.processing_end_time = now()
            db_entry.save()
        
        # Clean up any temporary files created during processing
        for temp_file in temp_files:
            try:
                if os.path.exists(temp_file):
                    os.remove(temp_file)
                    ProcessingLog.objects.create(
                        audio_file=original_audio,
                        message=f"Cleaned up temporary file: {temp_file}",
                        level="INFO"
                    )
            except Exception as cleanup_error:
                ProcessingLog.objects.create(
                    audio_file=original_audio,
                    message=f"Failed to clean up temporary file {temp_file}: {str(cleanup_error)}",
                    level="WARNING"
                )
        
        return False

def get_pending_audio_files():
    """
    Get all audio files that have been uploaded but not yet processed
    """
    from .models import OriginalAudioFile, Database
    return OriginalAudioFile.objects.filter(database_entry__status='Pending')

def process_pending_audio_files():
    """
    Process all pending audio files one by one
    Returns a tuple of (processed_count, failed_count)
    """
    # Import here to avoid circular imports
    from .tasks import process_pending_audio_files_batch
    return process_pending_audio_files_batch()

def get_processing_status():
    """
    Get the current processing status counts
    """
    from .models import OriginalAudioFile, Database
    from .tasks import get_processor_status
    
    # Get all audio files
    total_files = OriginalAudioFile.objects.count()
    
    # Get counts by status
    pending_files = Database.objects.filter(status='Pending').count()
    processing_files = Database.objects.filter(status='Processing').count()
    processed_files = Database.objects.filter(status='Processed').count()
    failed_files = Database.objects.filter(status='Failed').count()
    
    # Calculate missing or untracked files (files without a Database entry)
    files_with_db_entries = Database.objects.count()
    untracked_files = max(0, total_files - files_with_db_entries)
    
    # Verify the counts add up correctly
    expected_total = pending_files + processing_files + processed_files + failed_files + untracked_files
    if expected_total != total_files:
        # If there's a mismatch, recount directly from the database to ensure accuracy
        pending_files = Database.objects.filter(status='Pending').count()
        processing_files = Database.objects.filter(status='Processing').count()
        processed_files = Database.objects.filter(status='Processed').count()
        failed_files = Database.objects.filter(status='Failed').count()
        untracked_files = max(0, total_files - (pending_files + processing_files + processed_files + failed_files))
    
    return {
        'total': total_files,
        'pending': pending_files,
        'processing': processing_files,
        'processed': processed_files,
        'failed': failed_files,
        'untracked': untracked_files,
        'processor_status': get_processor_status()
    }

def search_zoo(zoo_name):
    """
    Get files by zoo name
    """
    try:
        # get zoo object
        zoo = Zoo.objects.get(zoo_name=zoo_name)
        
        # search by that object
        audio_files = OriginalAudioFile.objects.filter(zoo=zoo)
        
        return audio_files
    except ObjectDoesNotExist:
        print(f"No zoo found with name: {zoo_name}")
        return None

def search_animal(animal_type):
    """
    Get files by animal type
    """
    try:
        # search by animal type
        audio_files = OriginalAudioFile.objects.filter(animal_type=animal_type)
        
        return audio_files
    except ObjectDoesNotExist:
        print(f"No audio files found for animal type: {animal_type}")
        return None

def handle_duplicate_file(new_file, animal_type, zoo=None, uploaded_by=None):
    """
    Handles duplicate file detection and replacement.
    
    Args:
        new_file: The uploaded file object
        animal_type: Animal type for the file (species_name in lowercase with spaces replaced by underscores)
        zoo: Zoo object (optional)
        uploaded_by: User profile who uploaded the file (optional)
        
    Returns:
        tuple: (original_audio, is_duplicate, replaced_file_id)
            - original_audio: The saved OriginalAudioFile instance
            - is_duplicate: Boolean indicating if this was a duplicate
            - replaced_file_id: ID of the file that was replaced (if any)
    """
    from django.db import transaction
    import os
    from .models import AnimalTable
    
    # Find the corresponding animal in AnimalTable
    animal = None
    try:
        # Convert animal_type format (e.g., 'amur_leopard') to proper case for lookup (e.g., 'Amur Leopard')
        species_name = ' '.join(word.capitalize() for word in animal_type.split('_'))
        animal = AnimalTable.objects.filter(species_name__iexact=species_name).first()
    except Exception as e:
        # If there's an error finding the animal, log it but continue
        logging.error(f"Error finding animal for type {animal_type}: {str(e)}")
    
    # Check if a file with the same name already exists
    existing_file = OriginalAudioFile.objects.filter(audio_file_name=new_file.name).first()
    
    if existing_file:
        # We found a duplicate
        replaced_file_id = existing_file.file_id
        
        try:
            with transaction.atomic():
                # Log the duplicate detection with a more visible WARNING level
                ProcessingLog.objects.create(
                    audio_file=existing_file,
                    timestamp=now(),
                    level='WARNING',
                    message=f'DUPLICATE FILE DETECTED: {new_file.name} - Replacing previous version with new upload.'
                )
                
                # Get the old file path before deleting
                old_file_path = existing_file.audio_file.path if existing_file.audio_file else None
                
                # Delete related database entries
                # First, get the database entry to delete it properly
                db_entries = Database.objects.filter(audio_file=existing_file)
                for entry in db_entries:
                    entry.delete()
                
                # Delete any spectrograms associated with this file
                Spectrogram.objects.filter(audio_file=existing_file).delete()
                
                # Delete any detected noise files associated with this file
                DetectedNoiseAudioFile.objects.filter(original_file=existing_file).delete()
                
                # Delete the physical file if it exists
                if old_file_path and os.path.exists(old_file_path):
                    try:
                        os.remove(old_file_path)
                    except (PermissionError, OSError) as e:
                        # Log error but continue with replacement
                        ProcessingLog.objects.create(
                            audio_file=existing_file,
                            timestamp=now(),
                            level='ERROR',
                            message=f'Error deleting old file: {str(e)}'
                        )
                
                # Update the existing record with new file
                existing_file.audio_file = new_file
                existing_file.animal_type = animal_type
                # Associate with the animal from AnimalTable if found
                if animal:
                    existing_file.animal = animal
                if zoo:
                    existing_file.zoo = zoo
                if uploaded_by:
                    existing_file.uploaded_by = uploaded_by
                existing_file.upload_date = now()
                existing_file.recording_date = None  # Will be extracted from filename later
                existing_file.file_size_mb = None  # Will be calculated on save
                existing_file.save()
                
                # Create a new database entry with Pending status
                db_entry = Database(
                    audio_file=existing_file,
                    status='Pending'
                )
                db_entry.save()
                
                # Create a log entry for the replacement with a more visible SUCCESS level
                ProcessingLog.objects.create(
                    audio_file=existing_file,
                    timestamp=now(),
                    level='SUCCESS',
                    message=f'DUPLICATE REPLACED: {new_file.name} - File successfully replaced with new upload.'
                )
                
                return existing_file, True, replaced_file_id
                
        except Exception as e:
            # Log the error
            if existing_file:
                ProcessingLog.objects.create(
                    audio_file=existing_file,
                    timestamp=now(),
                    level='ERROR',
                    message=f'Error replacing duplicate file: {str(e)}'
                )
            # If replacement fails, we'll create a new file instead
            pass
    
    # No duplicate found or replacement failed, create a new file
    original_audio = OriginalAudioFile(
        audio_file=new_file,
        audio_file_name=new_file.name,
        animal_type=animal_type,
        zoo=zoo,
        upload_date=now(),
        uploaded_by=uploaded_by,
        animal=animal  # Associate with the animal from AnimalTable if found
    )
    original_audio.save()
    
    return original_audio, False, None


def process_audio(file_path, original_audio):
    """
    Process the uploaded audio file and store saw call timeframes.
    Uses improved STFT-based detection to accurately identify and log saw calls.
    """
    temp_files = []  # Track temporary files for cleanup
    
    try:
        # Update the database status to Processing
        db_entry = Database.objects.filter(audio_file=original_audio).first()
        if db_entry:
            db_entry.status = 'Processing'
            db_entry.processing_start_time = now()
            db_entry.save()
        
        # Load the audio file
        try:
            sample_rate, audio_data = wav.read(file_path)
            
            # Convert to mono if stereo
            if len(audio_data.shape) > 1:
                audio_data = np.mean(audio_data, axis=1).astype(audio_data.dtype)
            
            # Log successful audio loading
            ProcessingLog.objects.create(
                audio_file=original_audio,
                timestamp=now(),
                level='INFO',
                message=f'Audio file loaded successfully: {sample_rate}Hz, {len(audio_data)} samples'
            )
        except Exception as e:
            # Log error and raise to be caught by outer try/except
            ProcessingLog.objects.create(
                audio_file=original_audio,
                timestamp=now(),
                level='ERROR',
                message=f'Error loading audio file: {str(e)}'
            )
            raise
        
        # Detect saw calls
        saw_calls = detect_saw_calls(audio_data, sample_rate, animal_type=original_audio.animal_type)
        
        # Log detection results
        ProcessingLog.objects.create(
            audio_file=original_audio,
            timestamp=now(),
            level='INFO',
            message=f'Detected {len(saw_calls)} saw calls in the audio'
        )
        
        # Save each detected saw call as a separate audio file
        for i, call in enumerate(saw_calls):
            try:
                # Calculate start and end samples
                start_seconds = call['start_seconds']
                end_seconds = call['end_seconds']
                start_sample = int(start_seconds * sample_rate)
                end_sample = int(end_seconds * sample_rate)
                
                # Extract the audio segment
                segment = audio_data[start_sample:end_sample]
                
                # Create a temporary file for the segment
                temp_dir = os.path.join(settings.MEDIA_ROOT, 'temp')
                ensure_directory_exists(temp_dir)
                temp_file = os.path.join(temp_dir, f'temp_segment_{i}.wav')
                temp_files.append(temp_file)  # Track for cleanup
                
                # Save the segment to a temporary WAV file
                wav.write(temp_file, sample_rate, segment)
                
                # Create a filename for the segment
                segment_filename = f'{os.path.splitext(original_audio.audio_file_name)[0]}_segment_{i+1}.wav'
                
                # Create a DetectedNoiseAudioFile instance
                with open(temp_file, 'rb') as f:
                    # Convert start and end time strings to time objects if they're not already
                    if isinstance(call['start'], str):
                        start_time = datetime.strptime(call['start'], '%H:%M:%S').time()
                    else:
                        start_time = call['start']
                        
                    if isinstance(call['end'], str):
                        end_time = datetime.strptime(call['end'], '%H:%M:%S').time()
                    else:
                        end_time = call['end']
                    
                    # Calculate file size in MB
                    file_size_mb = os.path.getsize(temp_file) / (1024 * 1024)
                    
                    # Save the segment to a permanent location
                    segment_dir = os.path.join(settings.MEDIA_ROOT, 'detected_noises')
                    ensure_directory_exists(segment_dir)
                    permanent_segment_path = os.path.join(segment_dir, segment_filename)
                    
                    # Copy the file to the permanent location
                    shutil.copy2(temp_file, permanent_segment_path)
                    
                    # Create relative path for database storage
                    relative_path = os.path.join('detected_noises', segment_filename).replace('\\', '/')
                    
                    # Create the DetectedNoiseAudioFile instance
                    noise_file = DetectedNoiseAudioFile(
                        original_file=original_audio,
                        detected_noise_file_path=relative_path,
                        start_time=start_time,
                        end_time=end_time,
                        # Use saw_count and saw_call_count for the impulse count
                        saw_count=call['impulse_count'],
                        saw_call_count=1,  # Each detection is one call
                        frequency=call['frequency'],
                        magnitude=call['magnitude'],
                        file_size_mb=file_size_mb
                    )
                    noise_file.save()
                
                # Generate spectrogram for the segment using our new function
                try:
                    # First save the segment to a temporary file that can be read by our spectrogram function
                    segment_file = os.path.join(temp_dir, f'segment_for_spec_{i}.wav')
                    temp_files.append(segment_file)  # Track for cleanup
                    wav.write(segment_file, sample_rate, segment)
                    
                    # Generate mel spectrogram for the segment
                    segment_spectrogram = generate_spectrogram(
                        segment_file, 
                        original_audio, 
                        spectrogram_type='mel'
                    )
                    
                    # Associate the spectrogram with the noise file
                    if segment_spectrogram:
                        segment_spectrogram.noise_file = noise_file
                        segment_spectrogram.is_full_audio = False
                        segment_spectrogram.save()
                        
                        # Log successful spectrogram generation
                        ProcessingLog.objects.create(
                            audio_file=original_audio,
                            timestamp=now(),
                            level='INFO',
                            message=f'Spectrogram generated for segment {i+1} ({call["start"]} - {call["end"]})'
                        )
                    
                except Exception as e:
                    # Log error but continue processing
                    ProcessingLog.objects.create(
                        audio_file=original_audio,
                        timestamp=now(),
                        level='WARNING',
                        message=f'Error generating spectrogram for segment {i+1}: {str(e)}'
                    )
            
            except Exception as e:
                # Log error but continue with other segments
                ProcessingLog.objects.create(
                    audio_file=original_audio,
                    timestamp=now(),
                    level='WARNING',
                    message=f'Error processing segment {i+1}: {str(e)}'
                )
        
        # Generate spectrograms for the full audio using our new function
        try:
            # Generate mel spectrogram (good for general audio analysis)
            mel_spectrogram = generate_spectrogram(file_path, original_audio, spectrogram_type='mel')
            
            # Generate linear spectrogram (standard STFT)
            linear_spectrogram = generate_spectrogram(file_path, original_audio, spectrogram_type='linear')
            
            # For very long audio, log that we're using the optimized spectrogram function
            if len(audio_data) > 10000000:  # 10 million samples
                ProcessingLog.objects.create(
                    audio_file=original_audio,
                    timestamp=now(),
                    level='INFO',
                    message='Using optimized spectrogram generation for large audio file'
                )
            
            # Log successful spectrogram generation
            ProcessingLog.objects.create(
                audio_file=original_audio,
                timestamp=now(),
                level='INFO',
                message=f'Multiple spectrograms generated successfully for full audio'
            )
            
        except Exception as e:
            # Log error but continue processing
            ProcessingLog.objects.create(
                audio_file=original_audio,
                timestamp=now(),
                level='WARNING',
                message=f'Error generating full audio spectrogram: {str(e)}'
            )
        
        # Generate Excel report
        try:
            excel_path = generate_excel_report(original_audio, saw_calls)
            temp_files.append(excel_path)  # Track the Excel file for potential cleanup
            
            # Log Excel report generation
            ProcessingLog.objects.create(
                audio_file=original_audio,
                timestamp=now(),
                level='INFO',
                message=f'Excel report generated successfully: {excel_path}'
            )
        except Exception as e:
            # Log error but continue processing
            ProcessingLog.objects.create(
                audio_file=original_audio,
                timestamp=now(),
                level='ERROR',
                message=f'Error generating Excel report: {str(e)}'
            )
        
        # Update database status to Processed
        if db_entry:
            db_entry.status = 'Processed'
            db_entry.processing_end_time = now()
            db_entry.save()
        
        # Log successful processing
        ProcessingLog.objects.create(
            audio_file=original_audio,
            timestamp=now(),
            level='SUCCESS',
            message=f'Audio processing completed successfully. {len(saw_calls)} saw calls detected.'
        )
        
        return True
    
    except Exception as e:
        # Log the error
        ProcessingLog.objects.create(
            audio_file=original_audio,
            timestamp=now(),
            level='ERROR',
            message=f'Error processing audio file: {str(e)}'
        )
        
        # Update database status to Failed
        if db_entry:
            db_entry.status = 'Failed'
            db_entry.processing_end_time = now()
            db_entry.save()
        
        # Clean up any temporary files created during processing
        for temp_file in temp_files:
            try:
                if os.path.exists(temp_file):
                    os.remove(temp_file)
                    ProcessingLog.objects.create(
                        audio_file=original_audio,
                        timestamp=now(),
                        level='INFO',
                        message=f'Cleaned up temporary file: {temp_file}'
                    )
            except Exception as cleanup_error:
                ProcessingLog.objects.create(
                    audio_file=original_audio,
                    timestamp=now(),
                    level='WARNING',
                    message=f'Failed to clean up temporary file {temp_file}: {str(cleanup_error)}'
                )
        
        return False

def advanced_search_audio(search_params):
    """
    Advanced search function that combines multiple search criteria with optimized performance
    
    Args:
        search_params: Dictionary containing search parameters
            - upload_date_start: Start date for upload date range (datetime)
            - upload_date_end: End date for upload date range (datetime)
            - recording_date_start: Start date for recording date range (datetime)
            - recording_date_end: End date for recording date range (datetime)
            - device_id: Device ID string to search for in filename
            - animal_type: Animal type to filter by
            - zoo: Zoo ID to filter by
            - search_query: General search term for filename
            - limit: Maximum number of results to return (int, default: None)
            - offset: Number of results to skip (int, default: 0)
    
    Returns:
        QuerySet of OriginalAudioFile objects matching the criteria
    """
    # Validate search parameters
    if not isinstance(search_params, dict):
        raise ValueError("Search parameters must be provided as a dictionary")
    
    # Start with a base query that includes related objects to optimize performance
    # Using select_related for ForeignKey relationships to reduce database queries
    audio_files = OriginalAudioFile.objects.select_related(
        'zoo', 'uploaded_by', 'animal'
    ).prefetch_related(
        'detected_noises', 'database_entry', 'spectrograms'
    )
    
    # Apply filters based on provided parameters
    if search_params.get('search_query'):
        query = search_params['search_query'].strip()
        if query:
            # Use Q objects for more complex queries
            from django.db.models import Q
            audio_files = audio_files.filter(
                Q(audio_file_name__icontains=query) | 
                Q(animal_type__icontains=query)
            )
    
    if search_params.get('animal_type'):
        animal_type = search_params['animal_type']
        if isinstance(animal_type, str) and animal_type.strip():
            audio_files = audio_files.filter(animal_type=animal_type.strip())
    
    if search_params.get('zoo'):
        try:
            # Use the correct primary key field (zoo_id instead of id)
            zoo_id = int(search_params['zoo'])
            audio_files = audio_files.filter(zoo__zoo_id=zoo_id)
        except (ValueError, TypeError):
            # If zoo ID is not a valid integer, try string matching on zoo name
            zoo_name = str(search_params['zoo']).strip()
            if zoo_name:
                audio_files = audio_files.filter(zoo__zoo_name__icontains=zoo_name)
    
    if search_params.get('device_id'):
        device_id = str(search_params['device_id']).strip()
        if device_id:
            audio_files = audio_files.filter(audio_file_name__icontains=device_id)
    
    # Date range filters with validation
    from django.utils.timezone import now
    from datetime import datetime, timedelta
    
    # Helper function to validate date parameters
    def validate_date(date_param):
        if isinstance(date_param, (datetime, str)):
            if isinstance(date_param, str):
                try:
                    # Try to parse the date string
                    return datetime.strptime(date_param, '%Y-%m-%d')
                except ValueError:
                    return None
            return date_param
        return None
    
    # Upload date range
    upload_date_start = validate_date(search_params.get('upload_date_start'))
    if upload_date_start:
        audio_files = audio_files.filter(upload_date__gte=upload_date_start)
    
    upload_date_end = validate_date(search_params.get('upload_date_end'))
    if upload_date_end:
        # Add a day to include the end date fully
        upload_date_end = upload_date_end + timedelta(days=1)
        audio_files = audio_files.filter(upload_date__lt=upload_date_end)
    
    # Recording date range
    recording_date_start = validate_date(search_params.get('recording_date_start'))
    if recording_date_start:
        audio_files = audio_files.filter(recording_date__gte=recording_date_start)
    
    recording_date_end = validate_date(search_params.get('recording_date_end'))
    if recording_date_end:
        # Add a day to include the end date fully
        recording_date_end = recording_date_end + timedelta(days=1)
        audio_files = audio_files.filter(recording_date__lt=recording_date_end)
    
    # Add status filter if provided
    if search_params.get('status'):
        status = str(search_params['status']).strip()
        if status in ['Pending', 'Processing', 'Processed', 'Failed']:
            audio_files = audio_files.filter(database_entry__status=status)
    
    # Order by upload date (most recent first)
    audio_files = audio_files.order_by('-upload_date')
    
    # Apply pagination if specified
    try:
        limit = int(search_params.get('limit', 0))
        offset = int(search_params.get('offset', 0))
        
        if offset > 0:
            audio_files = audio_files[offset:]
        
        if limit > 0:
            audio_files = audio_files[:limit]
    except (ValueError, TypeError):
        # If pagination parameters are invalid, ignore them
        pass
    
    # Use distinct to avoid duplicate results
    return audio_files.distinct()